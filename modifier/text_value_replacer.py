import json
import logging
from typing import Any
import openpyxl
from constants import TEXTS_JSON, INPUT_XLXS

logger = logging.getLogger(__name__)

class TextValueReplacer:
    """
    Reads an Excel workbook and uses its cell coordinates as a lookup table
    to replace matching "value" fields inside a texts.json file.
    """

    def __init__(
        self,
        excel_path: str = INPUT_XLXS,
        json_path:  str = TEXTS_JSON
    ) -> None:
        self.excel_path = excel_path
        self.json_path  = json_path

    # ── public ─────────────────────────────────────────────────────────────────

    def run(self) -> list[dict[str, Any]]:
        """
        Full pipeline:
          1. Build the id→value map from Excel.
          2. Load texts.json.
          3. Replace matching values.
          4. Write the updated JSON back to disk.
          5. Return the updated list.
        """
        mapping   = self._build_mapping()
        texts     = self._load_json()
        updated   = self._apply_mapping(texts, mapping)
        self._save_json(updated)
        return updated

    def build_mapping(self) -> dict[str, str]:
        """Public accessor — returns the Excel id→value map without touching JSON."""
        return self._build_mapping()

    # ── private ────────────────────────────────────────────────────────────────

    def _build_mapping(self) -> dict[str, str]:
        """
        Scan every cell in every sheet of the workbook.
        Cell coordinate (e.g. "A1") → str(cell value).
        Cells with no value are skipped.
        """
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path, data_only=True)

        if "Sheet1" not in wb.sheetnames:
            raise ValueError("Workbook does not contain a sheet named 'Sheet1'.")

        # Guard: raise an error if any sheet other than Sheet1 contains data
        for sheet in wb.worksheets:
            if sheet.title == "Sheet1":
                continue
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        raise ValueError(
                            f"Unexpected data found in sheet '{sheet.title}' at cell "
                            f"{cell.coordinate} — only Sheet1 is allowed."
                        )

        mapping: dict[str, str] = {}
        for row in wb["Sheet1"].iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_id    = cell.coordinate          # e.g. "A1"
                cell_value = str(cell.value).strip()
                if cell_value:
                    mapping[cell_id] = cell_value

        logger.info("Excel mapping built — %d entries: %s", len(mapping), mapping)
        return mapping

    def _load_json(self) -> list[dict[str, Any]]:
        if not self.json_path.exists():
            raise FileNotFoundError(f"JSON file not found: {self.json_path}")

        with self.json_path.open(encoding="utf-8") as fh:
            data = json.load(fh)

        if not isinstance(data, list):
            raise ValueError(f"Expected a JSON array in {self.json_path}, got {type(data).__name__}")

        return data

    def _apply_mapping(
        self,
        texts:   list[dict[str, Any]],
        mapping: dict[str, str],
    ) -> list[dict[str, Any]]:
        """
        Iterate over every text entity.  If its "value" matches a key in the
        Excel mapping, replace it with the mapped string.
        """
        replaced = 0

        for entry in texts:
            original = entry.get("value", "")
            if original in mapping:
                new_value      = mapping[original]
                entry["value"] = new_value
                replaced      += 1
                logger.info("Replaced %r → %r", original, new_value)
            else:
                logger.debug("No match for %r — left unchanged", original)

        logger.info("Done — %d / %d entries replaced.", replaced, len(texts))
        return texts

    def _save_json(self, texts: list[dict[str, Any]]) -> None:
        self.json_path.parent.mkdir(parents=True, exist_ok=True)
        with self.json_path.open("w", encoding="utf-8") as fh:
            json.dump(texts, fh, indent=2, ensure_ascii=False)
        logger.info("Updated JSON written to %s", self.json_path)


# ── CLI entry point ────────────────────────────────────────────────────────────

def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s  %(message)s",
    )

    replacer = TextValueReplacer()

    try:
        updated = replacer.run()
    except FileNotFoundError as exc:
        logger.error(exc)
        raise SystemExit(1)

    print(f"\nDone — {len(updated)} text entities processed.")
    print("Updated values:")
    for entry in updated:
        print(f"  {entry['value']!r}  (layer: {entry.get('layer', '?')})")


if __name__ == "__main__":
    main()